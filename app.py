"""
AMS Factory Management Application  v4.0
FastAPI + SQLite — deployable on any Python server

Run locally:
    pip install -r requirements.txt
    python app.py  → http://localhost:8080

Deploy to cloud — set these environment variables:
    PORT         server port  (default 8080; Railway/Render set this automatically)
    DB_PATH      path to SQLite file (default: ./factory.db)
    REQUIRE_AUTH true/false   (default false — set true on cloud)
    FACTORY_USER login username  (default: factory)
    FACTORY_PASS login password  (default: ams2026 — CHANGE THIS)
"""
import sqlite3, json, os, base64, secrets
from pathlib import Path
from datetime import date as dt_date
from typing import Optional
import uvicorn
from fastapi import FastAPI, HTTPException, Body, Request, UploadFile, File
from fastapi.responses import HTMLResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

BASE   = Path(__file__).parent
DB     = os.getenv("DB_PATH", str(BASE / "factory.db"))
PORT   = int(os.getenv("PORT", 8080))
STATIC = BASE / "static"
STATIC.mkdir(exist_ok=True)

app = FastAPI(title="AMS Factory", version="4.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])

# ── OPTIONAL HTTP BASIC AUTH ───────────────────────────────────────────────────
@app.middleware("http")
async def auth_guard(request: Request, call_next):
    """Enabled only when REQUIRE_AUTH=true env var is set."""
    if os.getenv("REQUIRE_AUTH","false").lower() != "true":
        return await call_next(request)
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Basic "):
        try:
            user, pw = base64.b64decode(auth[6:]).decode("utf-8").split(":", 1)
            valid_u = os.getenv("FACTORY_USER", "factory")
            valid_p = os.getenv("FACTORY_PASS", "ams2026")
            if (secrets.compare_digest(user.encode(), valid_u.encode()) and
                    secrets.compare_digest(pw.encode(), valid_p.encode())):
                return await call_next(request)
        except Exception:
            pass
    return Response(status_code=401,
                    headers={"WWW-Authenticate": 'Basic realm="AMS Factory"'},
                    content="Unauthorized")

# ── DATABASE SCHEMA ────────────────────────────────────────────────────────────
SCHEMA = """
CREATE TABLE IF NOT EXISTS dl_modules (
    id        INTEGER PRIMARY KEY AUTOINCREMENT,
    account   TEXT NOT NULL,
    module    TEXT NOT NULL,
    routing   TEXT DEFAULT '',
    inc_mo    REAL DEFAULT 0, inc_h    REAL DEFAULT 1.5,
    sr_mo     REAL DEFAULT 0, sr_h     REAL DEFAULT 2.5,
    cr_mo     REAL DEFAULT 0, cr_h     REAL DEFAULT 8.0,
    nt_hrs    REAL DEFAULT 0,
    cplx      REAL DEFAULT 1.10,
    ai_pct    REAL DEFAULT 0.15,
    alloc_fte REAL DEFAULT 2.0,
    avail_pct REAL DEFAULT 0.85,
    skill_fit REAL DEFAULT 0.80,
    updated   TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS pod_modules (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    pod         TEXT NOT NULL,
    module      TEXT NOT NULL,
    module_lead TEXT DEFAULT 'TBD',
    inc_mo REAL DEFAULT 0, sr_mo REAL DEFAULT 0,
    cr_mo  REAL DEFAULT 0, nt_hrs REAL DEFAULT 0,
    cplx   REAL DEFAULT 1.15, ai_pct  REAL DEFAULT 0.18,
    alloc_fte REAL DEFAULT 2.0,
    updated TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS tower_skills (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    tower       TEXT NOT NULL,
    skill       TEXT NOT NULL,
    target_fte  REAL DEFAULT 0,
    ready_fte   REAL DEFAULT 0,
    stretch_fte REAL DEFAULT 0,
    shadow_fte  REAL DEFAULT 0,
    not_ready   REAL DEFAULT 0,
    demand_fte  REAL DEFAULT 0,
    cert_pct    REAL DEFAULT 0.6,
    updated TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS requirements (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    source      TEXT DEFAULT '',
    skill_module TEXT DEFAULT '',
    fte_req     REAL DEFAULT 0,
    by_month    TEXT DEFAULT '',
    priority    TEXT DEFAULT 'Medium',
    route_to    TEXT DEFAULT '',
    raised_by   TEXT DEFAULT '',
    date_raised TEXT DEFAULT '',
    status      TEXT DEFAULT 'Open',
    notes       TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS actions (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet       TEXT DEFAULT '',
    action_ref  TEXT DEFAULT '',
    action_date TEXT DEFAULT '',
    description TEXT DEFAULT '',
    owner       TEXT DEFAULT '',
    due_by      TEXT DEFAULT '',
    status      TEXT DEFAULT 'Open',
    outcome     TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS rebalancing (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    ref           TEXT DEFAULT '',
    resource      TEXT DEFAULT '',
    from_sheet    TEXT DEFAULT '',
    to_sheet      TEXT DEFAULT '',
    fte           REAL DEFAULT 0,
    reason        TEXT DEFAULT '',
    proposed_by   TEXT DEFAULT '',
    approved_by   TEXT DEFAULT '',
    decision_date TEXT DEFAULT '',
    status        TEXT DEFAULT 'Proposed',
    outcome       TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS resources (
    id              TEXT PRIMARY KEY,
    name            TEXT DEFAULT '',
    career_level    TEXT DEFAULT '',
    du              TEXT DEFAULT '',
    pod_tower       TEXT DEFAULT '',
    account         TEXT DEFAULT '',
    primary_skill   TEXT DEFAULT '',
    secondary_skill TEXT DEFAULT '',
    proficiency     INTEGER DEFAULT 3,
    cert            TEXT DEFAULT 'N',
    status          TEXT DEFAULT 'Active',
    alloc_fte       REAL DEFAULT 1.0,
    res_type        TEXT DEFAULT 'Human',
    skill_fit       REAL DEFAULT 0.8,
    notes           TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS settings (
    key     TEXT PRIMARY KEY,
    value   TEXT DEFAULT '',
    label   TEXT DEFAULT '',
    updated TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS monthly_tracker (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    account_id  TEXT NOT NULL,
    module_id   INTEGER NOT NULL,
    month       TEXT NOT NULL,
    demand_fte  REAL DEFAULT 0,
    cap_fte     REAL DEFAULT 0,
    notes       TEXT DEFAULT '',
    updated     TEXT DEFAULT '',
    UNIQUE(account_id, module_id, month)
);
CREATE TABLE IF NOT EXISTS ai_agents (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    pod_tower   TEXT NOT NULL,
    module_name TEXT DEFAULT 'All',
    agent_name  TEXT DEFAULT '',
    use_case    TEXT DEFAULT '',
    agent_type  TEXT DEFAULT 'Other',
    status      TEXT DEFAULT 'Planned',
    priority    TEXT DEFAULT 'Medium',
    fte_savings REAL DEFAULT 0,
    go_live     TEXT DEFAULT '',
    notes       TEXT DEFAULT '',
    updated     TEXT DEFAULT ''
);
"""

# ── SEED DATA ──────────────────────────────────────────────────────────────────
SEED_DL = [
    ("DL_ACCOUNT_A","ABAP/UI5","TWR_ABAP",800,1.2,90,2.5,8,12.0,0,1.15,0.20,22,0.85,0.80),
    ("DL_ACCOUNT_A","SD/GTS/TM","POD_OTC",400,1.2,45,2.5,5,8.0,0,1.15,0.15,14,0.85,0.80),
    ("DL_ACCOUNT_A","FI/CO","POD_R2R",350,1.2,35,2.5,3,8.0,0,1.10,0.15,12,0.85,0.80),
    ("DL_ACCOUNT_A","MM/eWM","POD_P2P",250,1.2,25,2.5,2,8.0,0,1.10,0.10,8,0.85,0.80),
    ("DL_ACCOUNT_A","BW/Reporting","POD_REPORTING",200,1.2,20,2.5,1,6.0,0,1.05,0.20,4,0.85,0.80),
    ("DL_ACCOUNT_B","FI/CO","POD_R2R",600,1.2,60,2.5,4,10.0,0,1.15,0.18,16,0.85,0.80),
    ("DL_ACCOUNT_B","ABAP/UI5","TWR_ABAP",500,1.2,50,2.5,5,10.0,0,1.10,0.20,18,0.85,0.80),
    ("DL_ACCOUNT_B","Non-SAP","TWR_ABAP",300,1.2,30,2.5,2,6.0,0,1.10,0.10,8,0.85,0.80),
    ("DL_ACCOUNT_B","SD/MM/EWM","POD_P2P",350,1.2,35,2.5,3,8.0,0,1.10,0.15,12,0.85,0.80),
    ("DL_ACCOUNT_B","BW/BTP","POD_REPORTING",200,1.2,20,2.5,2,6.0,0,1.05,0.22,6,0.85,0.80),
    ("DL_ACCOUNT_B","EHS","POD_EHS",150,1.0,15,2.0,1,6.0,0,1.05,0.10,5,0.85,0.80),
    ("DL_ACCOUNT_C","ABAP/UI5","TWR_ABAP",600,1.2,60,2.5,6,12.0,0,1.15,0.20,10,0.85,0.80),
    ("DL_ACCOUNT_C","FI/CO","POD_R2R",400,1.2,40,2.5,4,8.0,0,1.15,0.15,9,0.85,0.80),
    ("DL_ACCOUNT_C","SD/GTS/TM","POD_OTC",400,1.2,40,2.5,4,8.0,0,1.15,0.15,9,0.85,0.80),
    ("DL_ACCOUNT_C","EHS","POD_EHS",150,1.0,15,2.0,1,6.0,0,1.05,0.10,3,0.85,0.80),
    ("DL_ACCOUNT_C","MM/eWM","POD_P2P",200,1.2,20,2.5,2,8.0,0,1.10,0.10,5,0.85,0.80),
    ("DL_ACCOUNT_C","PM/QM/ePPDS","POD_MFG",200,1.2,20,2.5,3,8.0,0,1.10,0.10,5,0.85,0.80),
]
SEED_POD = [
    ("POD_R2R","FI Lead","TBD",0,0,0,0,1.15,0.18,14),
    ("POD_R2R","CO Lead","TBD",0,0,0,0,1.15,0.18,18),
    ("POD_R2R","Asset Acctg","TBD",0,0,0,0,1.10,0.15,8),
    ("POD_R2R","GL/AP/AR","TBD",0,0,0,0,1.10,0.15,9),
    ("POD_R2R","Tax/REFX","TBD",0,0,0,0,1.05,0.12,7),
    ("POD_OTC","SD Lead","TBD",0,0,0,0,1.15,0.18,20),
    ("POD_OTC","GTS Lead","TBD",0,0,0,0,1.15,0.18,12),
    ("POD_OTC","TM Lead","TBD",0,0,0,0,1.10,0.15,10),
    ("POD_OTC","Billing/Credit","TBD",0,0,0,0,1.10,0.12,8),
    ("POD_P2P","MM Lead","TBD",0,0,0,0,1.15,0.18,18),
    ("POD_P2P","EWM/WM Lead","TBD",0,0,0,0,1.15,0.18,16),
    ("POD_P2P","IM Lead","TBD",0,0,0,0,1.10,0.15,8),
    ("POD_MFG","PP/ePPDS Lead","TBD",0,0,0,0,1.15,0.18,16),
    ("POD_MFG","PM Lead","TBD",0,0,0,0,1.10,0.15,14),
    ("POD_MFG","QM Lead","TBD",0,0,0,0,1.10,0.15,13),
    ("POD_HCM","PA/OM Lead","TBD",0,0,0,0,1.10,0.15,4),
    ("POD_HCM","Payroll Lead","TBD",0,0,0,0,1.10,0.15,4),
    ("POD_HCM","Time Mgmt","TBD",0,0,0,0,1.05,0.12,3),
    ("POD_EHS","EHS Core","TBD",0,0,0,0,1.05,0.10,3),
    ("POD_EHS","Product Compliance","TBD",0,0,0,0,1.05,0.10,2),
    ("POD_REPORTING","BW/BI Lead","TBD",0,0,0,0,1.10,0.20,10),
    ("POD_REPORTING","BTP Analytics","TBD",0,0,0,0,1.10,0.20,7),
    ("POD_REPORTING","S/4 Analytics","TBD",0,0,0,0,1.08,0.18,6),
]
SEED_TWR = [
    ("TWR_ABAP","ABAP OO",60,39,12,6,3,36,0.70),
    ("TWR_ABAP","SmartForms/Adobe",35,23,7,3,2,21,0.65),
    ("TWR_ABAP","BADIs/Exits",40,26,8,4,2,24,0.70),
    ("TWR_ABAP","Fiori/UI5",42,27,8,5,2,25,0.60),
    ("TWR_ABAP","BSP/WebDynpro",18,12,4,1,1,11,0.65),
    ("TWR_ABAP","RFC/BAPI",19,12,4,2,1,11,0.70),
    ("TWR_BASIS","System Admin",8,5,2,1,0,5,0.75),
    ("TWR_BASIS","Transport Mgmt",5,3,1,1,0,3,0.70),
    ("TWR_BASIS","Performance",4,3,1,0,0,2,0.65),
    ("TWR_BASIS","BTP Infra",4,2,1,1,0,2,0.60),
    ("TWR_SECURITY","Role Design",10,6,2,1,1,6,0.70),
    ("TWR_SECURITY","GRC AC",8,5,2,1,0,5,0.70),
    ("TWR_SECURITY","GRC PC",6,4,1,1,0,4,0.65),
    ("TWR_SECURITY","SoD Analysis",5,3,1,1,0,3,0.65),
    ("TWR_INTEGRATION","PI/PO",8,5,2,1,0,5,0.70),
    ("TWR_INTEGRATION","CPI/BTP iFlow",9,6,2,1,0,5,0.65),
    ("TWR_INTEGRATION","API Management",5,3,1,1,0,3,0.60),
    ("TWR_INTEGRATION","EDI",4,3,1,0,0,2,0.70),
    ("TWR_AUTOMATION","ITSM Automation",3,2,1,0,0,2,0.70),
    ("TWR_AUTOMATION","AI Triage Agent",4,3,1,0,0,2,0.65),
    ("TWR_AUTOMATION","Agentic AI",3,2,1,0,0,2,0.60),
    ("TWR_AUTOMATION","RPA",2,1,1,0,0,1,0.65),
]
SEED_RES = [
    ("R001","Arun Sharma","Manager","DU-SAP","POD_R2R","Account Alpha","FI","CO",5,"Y","Active",0.8,"Human",0.8,"POD Lead R2R"),
    ("R002","Priya Nair","Manager","DU-SAP","POD_OTC","Account Alpha","SD","GTS",5,"Y","Active",0.8,"Human",0.8,"POD Lead OTC"),
    ("R003","Vikram Singh","Manager","DU-SAP","POD_P2P","Account Beta","MM","EWM",5,"Y","Active",0.8,"Human",0.8,"POD Lead P2P"),
    ("R004","Anjali Patel","Manager","DU-SAP","POD_MFG","Account Gamma","PP","PM",5,"Y","Active",0.8,"Human",0.8,"POD Lead MFG"),
    ("R005","Rahul Gupta","Manager","DU-SAP","POD_HCM","Account Beta","PA","Payroll",5,"Y","Active",0.8,"Human",0.8,"POD Lead HCM"),
    ("R006","Sonal Reddy","Manager","DU-SAP","POD_EHS","Account Gamma","EHS","Compliance",4,"Y","Active",0.8,"Human",0.8,"POD Lead EHS"),
    ("R007","Deepak Verma","Manager","DU-SAP","POD_REPORTING","Account Alpha","BW","BTP",4,"Y","Active",0.8,"Human",0.8,"POD Lead Reporting"),
    ("R008","Rohan Iyer","Sr Manager","DU-Tech","TWR_ABAP","All","ABAP OO","Fiori/UI5",5,"Y","Active",0.7,"Human",0.7,"Tower Lead ABAP"),
    ("R009","Kavita Menon","Manager","DU-Tech","TWR_SECURITY","All","GRC AC","Role Design",5,"Y","Active",0.7,"Human",0.7,"Tower Lead Security"),
    ("R010","Suresh Kumar","Manager","DU-Tech","TWR_INTEGRATION","All","CPI/BTP","API Mgmt",4,"Y","Active",0.7,"Human",0.7,"Tower Lead Integration"),
    ("R011","Ravi Bhat","Manager","DU-Tech","TWR_AUTOMATION","All","Agentic AI","ITSM",4,"Y","Active",0.7,"Human",0.7,"Tower Lead Automation"),
    ("AGT001","Incident Triage Agent","—","Automation","TWR_AUTOMATION","All","ITSM","AI Triage",4,"N","Active",1.0,"Agent",0.8,"L2 Operational"),
    ("AGT002","RCA Generator","—","Automation","TWR_AUTOMATION","All","Problem Mgmt","Analytics",4,"N","Active",1.0,"Agent",0.8,"L2 Operational"),
    ("AGT003","CR Sizing Agent","—","Automation","TWR_AUTOMATION","All","CR Estimation","Sizing",3,"N","Active",1.0,"Agent",0.7,"L1 Pilot"),
    ("AGT004","Status Report Agent","—","Automation","TWR_AUTOMATION","All","Reporting","Automation",4,"N","Active",1.0,"Agent",0.8,"L3 Operational"),
]
SEED_AGENTS = [
    # pod_tower, module_name, agent_name, use_case, agent_type, status, priority, fte_savings, go_live, notes
    ("POD_R2R","GL/AP/AR","GL Auto-Reconciliation","Auto-reconcile GL accounts end-of-month and flag discrepancies for human review","Reconciliation","Deployed","High",2.0,"2026-07","Processing 80% of standard reconciliations; 20% escalated to human"),
    ("POD_R2R","FI Lead","Incident Classifier","Classify FI/CO incidents from ticket text and auto-assign to resolver group","ITSM","Deployed","High",1.5,"2026-07","60% of FI incidents auto-closed at L1; saves ~1.5 FTE/month"),
    ("POD_OTC","SD Lead","Order Anomaly Detector","Detect pricing and credit anomalies on sales orders before dispatch","Analytics","Pilot","High",0.8,"2026-11","Catching 85% of pricing anomalies pre-dispatch in pilot accounts"),
    ("POD_P2P","MM Lead","Invoice Matching Agent","3-way PO/GR/Invoice match automation with exception routing","Reconciliation","Pilot","High",1.2,"2026-10","Targeting 70% straight-through processing for standard invoices"),
    ("POD_REPORTING","BW/BI Lead","Data Quality Monitor","Continuous monitoring of BW data flows with auto-alerting on failures","Analytics","Pilot","Medium",0.6,"2026-11","Monitoring 25 key data flows; 90% alerts resolved before business impact"),
    ("TWR_AUTOMATION","All","Incident Triage Agent","Auto-classify and route AMS incidents across all modules using LLM","ITSM","Deployed","High",3.5,"2026-07","Serving POD_R2R, POD_OTC, POD_P2P; L1 automation reduces volume by 40%"),
    ("TWR_AUTOMATION","All","RCA Generator","Automated root cause analysis from incident history patterns","Analytics","Deployed","High",1.0,"2026-07","Generates structured RCA in 2 mins vs 2-hour manual effort"),
    ("TWR_AUTOMATION","All","CR Sizing Agent","LLM-based change request effort estimation from natural language scope","Estimation","Pilot","Medium",1.2,"2026-10","Accuracy within 15% of actuals for standard functional CRs"),
    ("TWR_AUTOMATION","All","Status Report Agent","Auto-generate weekly status reports collating data from all PODs and accounts","Reporting","Deployed","Medium",1.0,"2026-07","Saves 4 hrs/week per account team; fully automated for L3 accounts"),
    ("TWR_ABAP","ABAP OO","Code Review Agent","Automated ABAP code review against coding standards and security patterns","Code Quality","Planned","Medium",0.5,"2026-12","Will reduce review cycle from 3 days to 4 hrs for standard developments"),
]

# ── HELPERS ────────────────────────────────────────────────────────────────────
def db():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c

def rag(gap: float, threshold: float = 2.0) -> str:
    return "red" if gap > threshold else "amber" if gap > 0 else "green"

def compute_dl(row: dict) -> dict:
    h = ((row["inc_mo"]*row["inc_h"] + row["sr_mo"]*row["sr_h"] +
          row["cr_mo"]*row["cr_h"] + row["nt_hrs"]) *
         row["cplx"] * (1 - row["ai_pct"]))
    d = round(h / 160, 1)
    c = round(row["alloc_fte"] * row["avail_pct"] * row["skill_fit"], 1)
    g = round(d - c, 1)
    return {**row, "demand_hrs": round(h,1), "demand_fte": d,
            "cap_fte": c, "gap_fte": g, "rag": rag(g)}

def compute_pod(row: dict) -> dict:
    h = ((row["inc_mo"]*1.5 + row["sr_mo"]*2.5 +
          row["cr_mo"]*8.0 + row["nt_hrs"]) *
         row["cplx"] * (1 - row["ai_pct"]))
    d = round(h / 160, 1)
    c = round(row["alloc_fte"] * 0.85, 1)
    g = round(d - c, 1)
    return {**row, "demand_fte": d, "cap_fte": c, "gap_fte": g, "rag": rag(g)}

def compute_twr(row: dict) -> dict:
    c   = row["ready_fte"]
    d   = row["demand_fte"]
    g   = round(d - c, 1)
    cov = round(c / row["target_fte"] * 100, 1) if row["target_fte"] else 0
    return {**row, "cap_fte": c, "gap_fte": g, "coverage_pct": cov, "rag": rag(g, 1.0)}

def acct_summary(aid: str, conn) -> dict:
    rows = [compute_dl(dict(r)) for r in conn.execute("SELECT * FROM dl_modules WHERE account=?", (aid,))]
    d = round(sum(r["demand_fte"] for r in rows), 1)
    c = round(sum(r["cap_fte"]    for r in rows), 1)
    g = round(d - c, 1)
    return {"id": aid, "demand_fte": d, "cap_fte": c, "gap_fte": g,
            "rag": rag(g), "modules": rows}

def pod_summary(pid: str, conn) -> dict:
    rows = [compute_pod(dict(r)) for r in conn.execute("SELECT * FROM pod_modules WHERE pod=?", (pid,))]
    d = round(sum(r["demand_fte"] for r in rows), 1)
    c = round(sum(r["cap_fte"]    for r in rows), 1)
    g = round(d - c, 1)
    return {"id": pid, "demand_fte": d, "cap_fte": c, "gap_fte": g,
            "rag": rag(g), "modules": rows}

def twr_summary(tid: str, conn) -> dict:
    rows = [compute_twr(dict(r)) for r in conn.execute("SELECT * FROM tower_skills WHERE tower=?", (tid,))]
    d = round(sum(r["demand_fte"] for r in rows), 1)
    c = round(sum(r["cap_fte"]    for r in rows), 1)
    g = round(d - c, 1)
    return {"id": tid, "demand_fte": d, "cap_fte": c, "gap_fte": g,
            "rag": rag(g, 1.0), "skills": rows}

def init_db():
    conn = db()
    conn.executescript(SCHEMA)
    conn.commit()
    for col, defn in [("signal_type","TEXT DEFAULT 'manual'"),
                      ("ack_by","TEXT DEFAULT ''"),
                      ("ack_date","TEXT DEFAULT ''")]:
        try:
            conn.execute(f"ALTER TABLE requirements ADD COLUMN {col} {defn}")
            conn.commit()
        except Exception:
            pass
    if conn.execute("SELECT COUNT(*) FROM dl_modules").fetchone()[0] == 0:
        conn.executemany("INSERT INTO dl_modules (account,module,routing,inc_mo,inc_h,sr_mo,sr_h,cr_mo,cr_h,nt_hrs,cplx,ai_pct,alloc_fte,avail_pct,skill_fit) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", SEED_DL)
        conn.executemany("INSERT INTO pod_modules (pod,module,module_lead,inc_mo,sr_mo,cr_mo,nt_hrs,cplx,ai_pct,alloc_fte) VALUES (?,?,?,?,?,?,?,?,?,?)", SEED_POD)
        conn.executemany("INSERT INTO tower_skills (tower,skill,target_fte,ready_fte,stretch_fte,shadow_fte,not_ready,demand_fte,cert_pct) VALUES (?,?,?,?,?,?,?,?,?)", SEED_TWR)
        conn.executemany("INSERT INTO resources (id,name,career_level,du,pod_tower,account,primary_skill,secondary_skill,proficiency,cert,status,alloc_fte,res_type,skill_fit,notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", SEED_RES)
        conn.commit()
    if conn.execute("SELECT COUNT(*) FROM ai_agents").fetchone()[0] == 0:
        conn.executemany("INSERT INTO ai_agents (pod_tower,module_name,agent_name,use_case,agent_type,status,priority,fte_savings,go_live,notes) VALUES (?,?,?,?,?,?,?,?,?,?)", SEED_AGENTS)
        conn.commit()
    if conn.execute("SELECT COUNT(*) FROM settings").fetchone()[0] == 0:
        SEED_SETTINGS = [
            ("contracted_fte", "199.5",  "Total Contracted FTE (commercial)"),
            ("pilot_start",    "2026-08-01", "Pilot Go-Live Date"),
            ("pilot_end",      "2027-03-31", "Pilot End Date"),
            ("scale_start",    "2027-04-01", "Full Scale Start"),
            ("ai_target_pct",  "0.20",   "AI Productivity Target (decimal)"),
            ("gap_red_threshold",   "2.0", "Gap RED threshold (FTE)"),
            ("gap_amber_threshold", "0.0", "Gap AMBER threshold (FTE)"),
        ]
        conn.executemany("INSERT INTO settings (key,value,label) VALUES (?,?,?)", SEED_SETTINGS)
        conn.commit()
    conn.close()

# ── ROUTES ─────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def root():
    p = STATIC / "index.html"
    return HTMLResponse(p.read_text(encoding="utf-8") if p.exists() else
                        "<h2>Run the server from the factory_app/ directory.</h2>")

@app.get("/api/summary")
def summary():
    conn = db()
    accts  = ["DL_ACCOUNT_A","DL_ACCOUNT_B","DL_ACCOUNT_C"]
    pods   = ["POD_R2R","POD_OTC","POD_P2P","POD_MFG","POD_HCM","POD_EHS","POD_REPORTING"]
    towers = ["TWR_ABAP","TWR_BASIS","TWR_SECURITY","TWR_INTEGRATION","TWR_AUTOMATION"]
    ad = [acct_summary(a, conn) for a in accts]
    pd_ = [pod_summary(p, conn) for p in pods]
    td = [twr_summary(t, conn) for t in towers]
    td_ = sum(a["demand_fte"] for a in ad)
    tc  = sum(a["cap_fte"]    for a in ad)
    agt = conn.execute("SELECT COALESCE(SUM(alloc_fte),0) FROM resources WHERE res_type='Agent' AND status='Active'").fetchone()[0]
    reqs    = [dict(r) for r in conn.execute("SELECT * FROM requirements WHERE (signal_type IS NULL OR signal_type='manual') ORDER BY id DESC LIMIT 10").fetchall()]
    signals = [dict(r) for r in conn.execute("SELECT * FROM requirements WHERE signal_type='demand_signal' AND status='Open' ORDER BY id DESC LIMIT 15").fetchall()]
    reb     = [dict(r) for r in conn.execute("SELECT * FROM rebalancing ORDER BY id DESC LIMIT 10").fetchall()]
    def cfg(key, default):
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row[0] if row else default
    contracted = float(cfg("contracted_fte", "199.5"))
    ai_pct     = float(cfg("ai_target_pct",  "0.22"))
    conn.close()
    return {"contracted_fte":contracted,"demand_fte":round(td_,1),"cap_fte":round(tc,1),
            "gap_fte":round(td_-tc,1),"agent_fte":round(agt,1),"ai_pct":ai_pct,
            "accounts":ad,"pods":pd_,"towers":td,
            "requirements":reqs,"demand_signals":signals,"rebalancing":reb}

@app.get("/api/account/{aid}")
def get_account(aid: str):
    conn = db()
    data = acct_summary(aid, conn)
    reqs = [dict(r) for r in conn.execute("SELECT * FROM requirements WHERE source=? ORDER BY id DESC", (aid,)).fetchall()]
    acts = [dict(r) for r in conn.execute("SELECT * FROM actions WHERE sheet=? ORDER BY id DESC LIMIT 15", (aid,)).fetchall()]
    t_mods = [dict(r) for r in conn.execute("SELECT id,module,routing FROM dl_modules WHERE account=? ORDER BY id", (aid,)).fetchall()]
    tracker = [dict(r) for r in conn.execute("SELECT * FROM monthly_tracker WHERE account_id=? ORDER BY module_id,month", (aid,)).fetchall()]
    conn.close()
    return {**data, "requirements": reqs, "actions": acts, "tracker_modules": t_mods, "tracker": tracker}

@app.patch("/api/module/{mid}")
def patch_module(mid: int, body: dict = Body(...)):
    allowed = {"inc_mo","inc_h","sr_mo","sr_h","cr_mo","cr_h","nt_hrs","cplx","ai_pct","alloc_fte","avail_pct","skill_fit","routing"}
    conn = db()
    if not conn.execute("SELECT id FROM dl_modules WHERE id=?", (mid,)).fetchone():
        raise HTTPException(404)
    old = dict(conn.execute("SELECT * FROM dl_modules WHERE id=?", (mid,)).fetchone())
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE dl_modules SET {sets}, updated=? WHERE id=?",
                     [*upd.values(), str(dt_date.today()), mid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM dl_modules WHERE id=?", (mid,)).fetchone())
    old_d = compute_dl(old)["demand_fte"]
    new_d = compute_dl(row)["demand_fte"]
    signal_fired = False
    if upd and old_d > 0 and abs(new_d - old_d) / old_d > 0.03:
        delta = round(new_d - old_d, 1)
        direction = "UP" if delta > 0 else "DOWN"
        pri = "High" if abs(delta) >= 1.0 else "Medium"
        conn.execute("""INSERT INTO requirements
            (source,skill_module,fte_req,by_month,priority,route_to,raised_by,date_raised,notes,signal_type,status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (row["account"], row["module"], abs(delta), str(dt_date.today()), pri,
             row["routing"], "System", str(dt_date.today()),
             f"Demand {direction} {abs(delta):.1f} FTE — {row['module']} ({row['account']}) | was {old_d}, now {new_d}",
             "demand_signal", "Open"))
        conn.commit()
        signal_fired = True
    conn.close()
    return {**compute_dl(row), "signal_fired": signal_fired}

@app.get("/api/pod/{pid}")
def get_pod(pid: str):
    conn = db()
    data = pod_summary(pid, conn)
    reqs = [dict(r) for r in conn.execute("SELECT * FROM requirements WHERE route_to=? ORDER BY id DESC", (pid,)).fetchall()]
    agents = [dict(r) for r in conn.execute("SELECT * FROM ai_agents WHERE pod_tower=? ORDER BY priority DESC,id", (pid,)).fetchall()]
    conn.close()
    return {**data, "inbound": reqs, "agents": agents}

@app.patch("/api/pod_module/{mid}")
def patch_pod_module(mid: int, body: dict = Body(...)):
    allowed = {"inc_mo","sr_mo","cr_mo","nt_hrs","cplx","ai_pct","alloc_fte","module_lead"}
    conn = db()
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE pod_modules SET {sets}, updated=? WHERE id=?",
                     [*upd.values(), str(dt_date.today()), mid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM pod_modules WHERE id=?", (mid,)).fetchone())
    conn.close()
    return compute_pod(row)

@app.get("/api/tower/{tid}")
def get_tower(tid: str):
    conn = db()
    data = twr_summary(tid, conn)
    reqs = [dict(r) for r in conn.execute("SELECT * FROM requirements WHERE route_to=? ORDER BY id DESC LIMIT 5", (tid,)).fetchall()]
    agents = [dict(r) for r in conn.execute("SELECT * FROM ai_agents WHERE pod_tower=? ORDER BY priority DESC,id", (tid,)).fetchall()]
    conn.close()
    return {**data, "inbound": reqs, "agents": agents}

@app.patch("/api/tower_skill/{sid}")
def patch_tower_skill(sid: int, body: dict = Body(...)):
    allowed = {"ready_fte","stretch_fte","shadow_fte","not_ready","demand_fte","target_fte","cert_pct"}
    conn = db()
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE tower_skills SET {sets}, updated=? WHERE id=?",
                     [*upd.values(), str(dt_date.today()), sid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM tower_skills WHERE id=?", (sid,)).fetchone())
    conn.close()
    return compute_twr(row)

@app.get("/api/requirements")
def get_requirements(route_to: str = "", source: str = ""):
    conn = db()
    if route_to:
        rows = conn.execute("SELECT * FROM requirements WHERE route_to=? ORDER BY id DESC", (route_to,)).fetchall()
    elif source:
        rows = conn.execute("SELECT * FROM requirements WHERE source=? ORDER BY id DESC", (source,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM requirements ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/requirements")
def add_requirement(body: dict = Body(...)):
    conn = db()
    conn.execute("INSERT INTO requirements (source,skill_module,fte_req,by_month,priority,route_to,raised_by,date_raised,notes) VALUES (?,?,?,?,?,?,?,?,?)",
        (body.get("source",""), body.get("skill_module",""), float(body.get("fte_req",0)),
         body.get("by_month",""), body.get("priority","Medium"), body.get("route_to",""),
         body.get("raised_by",""), str(dt_date.today()), body.get("notes","")))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = dict(conn.execute("SELECT * FROM requirements WHERE id=?", (nid,)).fetchone())
    conn.close()
    return row

@app.patch("/api/requirements/{rid}")
def update_req(rid: int, body: dict = Body(...)):
    allowed = {"status","notes","fte_req","by_month","priority","approved_by"}
    conn = db()
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE requirements SET {sets} WHERE id=?", [*upd.values(), rid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM requirements WHERE id=?", (rid,)).fetchone())
    conn.close()
    return row

@app.get("/api/rebalancing")
def get_rebalancing():
    conn = db(); rows = conn.execute("SELECT * FROM rebalancing ORDER BY id DESC").fetchall(); conn.close()
    return [dict(r) for r in rows]

@app.post("/api/rebalancing")
def add_rebalancing(body: dict = Body(...)):
    conn = db()
    conn.execute("INSERT INTO rebalancing (ref,resource,from_sheet,to_sheet,fte,reason,proposed_by,status) VALUES (?,?,?,?,?,?,?,'Proposed')",
        (body.get("ref",""), body.get("resource",""), body.get("from_sheet",""),
         body.get("to_sheet",""), float(body.get("fte",0)), body.get("reason",""), body.get("proposed_by","")))
    conn.commit(); conn.close()
    return {"ok": True}

@app.patch("/api/rebalancing/{rid}")
def update_rebalancing(rid: int, body: dict = Body(...)):
    allowed = {"status","approved_by","decision_date","outcome"}
    conn = db()
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE rebalancing SET {sets} WHERE id=?", [*upd.values(), rid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM rebalancing WHERE id=?", (rid,)).fetchone())
    conn.close()
    return row

@app.get("/api/resources")
def get_resources():
    conn = db()
    rows = conn.execute("SELECT * FROM resources ORDER BY res_type, pod_tower, name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/actions")
def add_action(body: dict = Body(...)):
    conn = db()
    conn.execute("INSERT INTO actions (sheet,action_ref,action_date,description,owner,due_by,status) VALUES (?,?,?,?,?,?,'Open')",
        (body.get("sheet",""), body.get("ref",""), str(dt_date.today()),
         body.get("description",""), body.get("owner",""), body.get("due_by","")))
    conn.commit(); conn.close()
    return {"ok": True}

@app.post("/api/resources")
def add_resource(body: dict = Body(...)):
    conn = db()
    rid = body.get("id","").strip().upper()
    if not rid:
        raise HTTPException(400, detail="id is required")
    if conn.execute("SELECT id FROM resources WHERE id=?", (rid,)).fetchone():
        raise HTTPException(409, detail="Resource ID already exists")
    conn.execute("""INSERT INTO resources
        (id,name,career_level,du,pod_tower,account,primary_skill,secondary_skill,
         proficiency,cert,status,alloc_fte,res_type,skill_fit,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (rid, body.get("name",""), body.get("career_level",""), body.get("du",""),
         body.get("pod_tower",""), body.get("account",""),
         body.get("primary_skill",""), body.get("secondary_skill",""),
         int(body.get("proficiency",3)), body.get("cert","N"),
         body.get("status","Active"), float(body.get("alloc_fte",1.0)),
         body.get("res_type","Human"), float(body.get("skill_fit",0.8)),
         body.get("notes","")))
    conn.commit()
    row = dict(conn.execute("SELECT * FROM resources WHERE id=?", (rid,)).fetchone())
    conn.close()
    return row

@app.patch("/api/resources/{rid}")
def update_resource(rid: str, body: dict = Body(...)):
    allowed = {"name","career_level","du","pod_tower","account","primary_skill",
               "secondary_skill","proficiency","cert","status","alloc_fte",
               "res_type","skill_fit","notes"}
    conn = db()
    if not conn.execute("SELECT id FROM resources WHERE id=?", (rid,)).fetchone():
        raise HTTPException(404)
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE resources SET {sets} WHERE id=?", [*upd.values(), rid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM resources WHERE id=?", (rid,)).fetchone())
    conn.close()
    return row

@app.delete("/api/resources/{rid}")
def delete_resource(rid: str):
    conn = db()
    if not conn.execute("SELECT id FROM resources WHERE id=?", (rid,)).fetchone():
        raise HTTPException(404)
    conn.execute("DELETE FROM resources WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/simulate")
def simulate(body: dict = Body(...)):
    """Compute demand/cap/gap without saving — for what-if analysis."""
    t = body.get("type","dl")
    row = body.get("row",{})
    if t == "dl":
        return compute_dl(row)
    elif t == "pod":
        return compute_pod(row)
    elif t == "tower":
        return compute_twr(row)
    raise HTTPException(400, detail="type must be dl, pod, or tower")

@app.get("/api/export")
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        conn = db()
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Factory Summary"
        hdr_fill = PatternFill("solid", fgColor="0D3B5E")
        hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        # Accounts
        data = summary()
        ws.append(["Account","Demand FTE","Cap FTE","Gap FTE","RAG"])
        for c in ws[1]: c.fill = hdr_fill; c.font = hdr_font
        for a in data["accounts"]:
            ws.append([a["id"], a["demand_fte"], a["cap_fte"], a["gap_fte"], a["rag"]])
        ws.append([])
        ws.append(["POD","Demand FTE","Cap FTE","Gap FTE","RAG"])
        for c in ws[ws.max_row]: c.fill = hdr_fill; c.font = hdr_font
        for p in data["pods"]:
            ws.append([p["id"], p["demand_fte"], p["cap_fte"], p["gap_fte"], p["rag"]])
        conn.close()
        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        return Response(buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=factory_export.xlsx"})
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/settings")
def get_settings():
    conn = db()
    rows = conn.execute("SELECT key,value,label FROM settings ORDER BY key").fetchall()
    conn.close()
    return {r["key"]: {"value": r["value"], "label": r["label"]} for r in rows}

@app.patch("/api/settings")
def update_settings(body: dict = Body(...)):
    conn = db()
    for k, v in body.items():
        conn.execute(
            "INSERT INTO settings (key,value,updated) VALUES (?,?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated=excluded.updated",
            (k, str(v), str(dt_date.today())))
    conn.commit()
    rows = conn.execute("SELECT key,value,label FROM settings ORDER BY key").fetchall()
    conn.close()
    return {r["key"]: {"value": r["value"], "label": r["label"]} for r in rows}

@app.post("/api/dl_modules")
def add_dl_module(body: dict = Body(...)):
    conn = db()
    conn.execute(
        "INSERT INTO dl_modules (account,module,routing,inc_mo,inc_h,sr_mo,sr_h,cr_mo,cr_h,nt_hrs,cplx,ai_pct,alloc_fte,avail_pct,skill_fit) VALUES (?,?,?,0,1.5,0,2.5,0,8.0,0,1.10,0.15,?,0.85,0.80)",
        (body.get("account",""), body.get("module","New Module"), body.get("routing",""),
         float(body.get("alloc_fte", 2.0))))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = dict(conn.execute("SELECT * FROM dl_modules WHERE id=?", (nid,)).fetchone())
    conn.close()
    return compute_dl(row)

@app.delete("/api/module/{mid}")
def delete_dl_module(mid: int):
    conn = db()
    conn.execute("DELETE FROM dl_modules WHERE id=?", (mid,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.post("/api/pod_modules")
def add_pod_module(body: dict = Body(...)):
    conn = db()
    conn.execute(
        "INSERT INTO pod_modules (pod,module,module_lead,inc_mo,sr_mo,cr_mo,nt_hrs,cplx,ai_pct,alloc_fte) VALUES (?,?,?,0,0,0,0,1.10,0.15,?)",
        (body.get("pod",""), body.get("module","New Module"), body.get("module_lead","TBD"),
         float(body.get("alloc_fte", 2.0))))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = dict(conn.execute("SELECT * FROM pod_modules WHERE id=?", (nid,)).fetchone())
    conn.close()
    return compute_pod(row)

@app.delete("/api/pod_module/{mid}")
def delete_pod_module(mid: int):
    conn = db()
    conn.execute("DELETE FROM pod_modules WHERE id=?", (mid,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.post("/api/tower_skills")
def add_tower_skill(body: dict = Body(...)):
    conn = db()
    conn.execute(
        "INSERT INTO tower_skills (tower,skill,target_fte,ready_fte,stretch_fte,shadow_fte,not_ready,demand_fte,cert_pct) VALUES (?,?,?,0,0,1,1,?,0.60)",
        (body.get("tower",""), body.get("skill","New Skill"),
         float(body.get("target_fte", 5)), float(body.get("demand_fte", 3))))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = dict(conn.execute("SELECT * FROM tower_skills WHERE id=?", (nid,)).fetchone())
    conn.close()
    return compute_twr(row)

@app.delete("/api/tower_skill/{sid}")
def delete_tower_skill(sid: int):
    conn = db()
    conn.execute("DELETE FROM tower_skills WHERE id=?", (sid,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.get("/api/pod/{pid}/resources")
def pod_resources(pid: str):
    conn = db()
    rows = conn.execute("SELECT * FROM resources WHERE pod_tower=? ORDER BY res_type,name", (pid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/tower/{tid}/resources")
def tower_resources(tid: str):
    conn = db()
    rows = conn.execute("SELECT * FROM resources WHERE pod_tower=? ORDER BY res_type,name", (tid,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/upload/excel")
async def upload_excel(section: str = "", file: UploadFile = File(...)):
    try:
        import openpyxl, io
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"ok": False, "error": "File has no data rows"}
        headers = [str(h).strip().lower().replace(" ","_") if h else "" for h in rows[0]]
        updated = 0
        conn = db()
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            mid = d.get("id") or d.get("module_id") or d.get("skill_id")
            if not mid:
                continue
            if section == "dl":
                allowed = {"inc_mo","inc_h","sr_mo","sr_h","cr_mo","cr_h","nt_hrs","cplx","ai_pct","alloc_fte","avail_pct","skill_fit"}
            elif section == "pod":
                allowed = {"inc_mo","sr_mo","cr_mo","nt_hrs","cplx","ai_pct","alloc_fte"}
            elif section == "tower":
                allowed = {"ready_fte","stretch_fte","shadow_fte","not_ready","demand_fte","target_fte","cert_pct"}
            else:
                continue
            upd = {k: float(v) for k, v in d.items() if k in allowed and v is not None}
            if upd:
                tbl = "dl_modules" if section == "dl" else ("pod_modules" if section == "pod" else "tower_skills")
                sets = ", ".join(f"{k}=?" for k in upd)
                conn.execute(f"UPDATE {tbl} SET {sets}, updated=? WHERE id=?",
                             [*upd.values(), str(dt_date.today()), int(mid)])
                updated += 1
        conn.commit(); conn.close()
        return {"ok": True, "updated": updated}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ── MONTHLY TRACKER ────────────────────────────────────────────────────────────
@app.post("/api/tracker")
def upsert_tracker(body: dict = Body(...)):
    conn = db()
    conn.execute("""INSERT INTO monthly_tracker (account_id,module_id,month,demand_fte,cap_fte,notes,updated)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(account_id,module_id,month) DO UPDATE SET
        demand_fte=excluded.demand_fte, cap_fte=excluded.cap_fte,
        notes=excluded.notes, updated=excluded.updated""",
        (body["account_id"], int(body["module_id"]), body["month"],
         float(body.get("demand_fte", 0)), float(body.get("cap_fte", 0)),
         body.get("notes", ""), str(dt_date.today())))
    conn.commit()
    row = dict(conn.execute(
        "SELECT * FROM monthly_tracker WHERE account_id=? AND module_id=? AND month=?",
        (body["account_id"], int(body["module_id"]), body["month"])).fetchone())
    conn.close()
    return row

# ── DEMAND SIGNALS ─────────────────────────────────────────────────────────────
@app.get("/api/demand_feed")
def demand_feed():
    conn = db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM requirements WHERE signal_type='demand_signal' ORDER BY id DESC LIMIT 30"
    ).fetchall()]
    conn.close()
    return rows

@app.patch("/api/requirements/{rid}/ack")
def ack_requirement(rid: int, body: dict = Body(...)):
    conn = db()
    conn.execute("UPDATE requirements SET status='Acknowledged',ack_by=?,ack_date=? WHERE id=?",
                 (body.get("ack_by", "Lead"), str(dt_date.today()), rid))
    conn.commit()
    req = dict(conn.execute("SELECT * FROM requirements WHERE id=?", (rid,)).fetchone())
    fte_updated_on = None
    # Propagate FTE change to the receiving POD or Tower module
    if req.get("signal_type") == "demand_signal":
        pod_tower = req.get("route_to", "")
        module_nm = req.get("skill_module", "")
        fte_delta = float(req.get("fte_req") or 0)
        if fte_delta > 0 and pod_tower.startswith("POD_"):
            modules = conn.execute(
                "SELECT * FROM pod_modules WHERE pod=? ORDER BY alloc_fte DESC", (pod_tower,)
            ).fetchall()
            best = None
            ml = module_nm.lower()
            for m in modules:
                md = dict(m)
                if ml in md["module"].lower() or md["module"].lower() in ml:
                    best = md; break
            if not best and modules:
                best = dict(modules[0])
            if best:
                conn.execute(
                    "UPDATE pod_modules SET alloc_fte=ROUND(alloc_fte+?,2),updated=? WHERE id=?",
                    (fte_delta, str(dt_date.today()), best["id"]))
                conn.commit()
                fte_updated_on = best["module"]
        elif fte_delta > 0 and pod_tower.startswith("TWR_"):
            skills = conn.execute(
                "SELECT * FROM tower_skills WHERE tower=? ORDER BY demand_fte DESC", (pod_tower,)
            ).fetchall()
            best = None
            ml = module_nm.lower()
            for s in skills:
                sd = dict(s)
                if ml in sd["skill"].lower() or sd["skill"].lower() in ml:
                    best = sd; break
            if not best and skills:
                best = dict(skills[0])
            if best:
                conn.execute(
                    "UPDATE tower_skills SET demand_fte=ROUND(demand_fte+?,2),updated=? WHERE id=?",
                    (fte_delta, str(dt_date.today()), best["id"]))
                conn.commit()
                fte_updated_on = best["skill"]
    conn.close()
    return {**req, "fte_updated_on": fte_updated_on}

@app.get("/api/signals/count")
def signal_counts():
    conn = db()
    rows = conn.execute("""SELECT route_to, COUNT(*) as cnt FROM requirements
        WHERE signal_type='demand_signal' AND status='Open' GROUP BY route_to""").fetchall()
    conn.close()
    return {r["route_to"]: r["cnt"] for r in rows}

# ── AI AGENTS ──────────────────────────────────────────────────────────────────
@app.get("/api/ai_agents")
def get_ai_agents(pod_tower: str = ""):
    conn = db()
    if pod_tower:
        rows = conn.execute("SELECT * FROM ai_agents WHERE pod_tower=? ORDER BY priority DESC,id", (pod_tower,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM ai_agents ORDER BY pod_tower,priority DESC,id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/ai_agents")
def add_ai_agent(body: dict = Body(...)):
    conn = db()
    conn.execute("""INSERT INTO ai_agents
        (pod_tower,module_name,agent_name,use_case,agent_type,status,priority,fte_savings,go_live,notes,updated)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (body.get("pod_tower",""), body.get("module_name","All"),
         body.get("agent_name",""), body.get("use_case",""),
         body.get("agent_type","Other"), body.get("status","Planned"),
         body.get("priority","Medium"), float(body.get("fte_savings",0)),
         body.get("go_live",""), body.get("notes",""), str(dt_date.today())))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    row = dict(conn.execute("SELECT * FROM ai_agents WHERE id=?", (nid,)).fetchone())
    conn.close()
    return row

@app.patch("/api/ai_agents/{aid}")
def update_ai_agent(aid: int, body: dict = Body(...)):
    allowed = {"module_name","agent_name","use_case","agent_type","status","priority","fte_savings","go_live","notes"}
    conn = db()
    upd = {k: v for k, v in body.items() if k in allowed}
    if upd:
        sets = ", ".join(f"{k}=?" for k in upd)
        conn.execute(f"UPDATE ai_agents SET {sets},updated=? WHERE id=?",
                     [*upd.values(), str(dt_date.today()), aid])
        conn.commit()
    row = dict(conn.execute("SELECT * FROM ai_agents WHERE id=?", (aid,)).fetchone())
    conn.close()
    return row

@app.delete("/api/ai_agents/{aid}")
def delete_ai_agent(aid: int):
    conn = db()
    conn.execute("DELETE FROM ai_agents WHERE id=?", (aid,))
    conn.commit(); conn.close()
    return {"ok": True}

if __name__ == "__main__":
    init_db()
    print("[OK] Database:", DB)
    print("[OK] AMS Factory App -- http://localhost:8080")
    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info")
